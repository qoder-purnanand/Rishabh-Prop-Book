import openpyxl
import sys

output_lines = []

def p(s=''):
    print(s)
    output_lines.append(str(s))

wb = openpyxl.load_workbook(r'c:\Users\admin\OneDrive - Qode\Research\Portfolio Review\Rishabh Prop Book\Rishabh Account Summary - 31.03.2026 - V2 (New).xlsx', data_only=True)

p('='*80)
p('ALL SHEET NAMES:')
p('='*80)
for name in wb.sheetnames:
    p('  - ' + name)
p()

# ---- OVERVIEW: first 10 rows of every sheet ----
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    total_rows = len(rows)
    p('='*80)
    p(f"SHEET: '{sheet_name}' | Total rows: {total_rows} | Max col: {ws.max_column}")
    p('='*80)
    p('First 10 rows:')
    for i, row in enumerate(rows[:10]):
        p(f'  Row {i+1}: {row}')
    p()

# ---- NETWORTH INPUTS: full content ----
p('='*80)
p("FULL CONTENT: 'Networth Inputs'")
p('='*80)
ws = wb['Networth Inputs']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows):
    if any(cell is not None for cell in row):
        p(f'  Row {i+1}: {row}')
p()

# ---- CASH AND MARGIN: full content ----
p('='*80)
p("FULL CONTENT: 'Cash and Margin'")
p('='*80)
ws = wb['Cash and Margin']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows):
    if any(cell is not None for cell in row):
        p(f'  Row {i+1}: {row}')
p()
p('--- Cash and Margin: Columns V-Z specifically (col indices 21-25, 0-based) ---')
for i, row in enumerate(rows):
    cols_vz = row[21:26] if len(row) > 21 else ()
    if any(c is not None for c in cols_vz):
        p(f'  Row {i+1} cols V-Z: {cols_vz}')
p()

# ---- MASTERSHEET: first 20 rows ----
p('='*80)
p("FIRST 20 ROWS: 'Mastersheet'")
p('='*80)
ws = wb['Mastersheet']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows[:20]):
    p(f'  Row {i+1}: {row}')
p()

# ---- STRATEGY ALLOCATION: first 20 rows ----
p('='*80)
p("FIRST 20 ROWS: 'Strategy Allocation'")
p('='*80)
ws = wb['Strategy Allocation']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows[:20]):
    p(f'  Row {i+1}: {row}')
p()

# ---- EQUITY HOLDINGS: first 10 rows ----
p('='*80)
p("FIRST 10 ROWS: 'Equity Holdings'")
p('='*80)
ws = wb['Equity Holdings']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows[:10]):
    p(f'  Row {i+1}: {row}')
p()

# ---- MUTUAL FUND HOLDINGS: first 10 rows ----
p('='*80)
p("FIRST 10 ROWS: 'Mutual Fund Holdings'")
p('='*80)
ws = wb['Mutual Fund Holdings']
rows = list(ws.iter_rows(values_only=True))
p(f"Total rows: {len(rows)} | Max col: {ws.max_column}")
for i, row in enumerate(rows[:10]):
    p(f'  Row {i+1}: {row}')
p()

# Write to output file
with open(r'c:\Users\admin\OneDrive - Qode\Research\Portfolio Review\Rishabh Prop Book\excel_analysis_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print("DONE - output written to excel_analysis_output.txt")
