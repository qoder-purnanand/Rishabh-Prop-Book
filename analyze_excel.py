import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\admin\OneDrive - Qode\Research\Portfolio Review\Rishabh Prop Book\Rishabh Account Summary - 31.03.2026 - V2 (New).xlsx', data_only=True)

print('='*80)
print('ALL SHEET NAMES:')
print('='*80)
for name in wb.sheetnames:
    print('  -', name)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    total_rows = len(rows)
    print('='*80)
    print(f"SHEET: '{sheet_name}' | Total rows: {total_rows} | Max col: {ws.max_column}")
    print('='*80)
    print('First 10 rows:')
    for i, row in enumerate(rows[:10]):
        print(f'  Row {i+1}: {row}')
    print()
